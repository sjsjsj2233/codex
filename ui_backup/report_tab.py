import sys
import os
import logging
import csv
import re


# PyQt5 라이브러리
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTextEdit, QTableWidget, 
    QHBoxLayout, QGroupBox, QCheckBox, QPushButton, 
    QFileDialog, QTableWidgetItem, QHeaderView, QSplitter, 
    QMessageBox
)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt

# 엑셀 라이브러리
from openpyxl import Workbook

class EnhancedInspectionReportGenerator(QWidget):
    def __init__(self):
        super().__init__()
        self.device_data = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 상단 도구 모음
        tools_layout = QHBoxLayout()
        
        # 파일 로드 섹션
        load_group = QGroupBox("데이터 소스")
        load_layout = QHBoxLayout()
        
        load_button = QPushButton("파일 불러오기")
        load_button.setIcon(QIcon("icons/open.png"))
        load_button.clicked.connect(self.load_txt_files)
        
        self.file_count_label = QLabel("로드된 파일: 0개")
        
        clear_button = QPushButton("초기화")
        clear_button.setIcon(QIcon("icons/clear.png"))
        clear_button.clicked.connect(self.clear_data)
        
        load_layout.addWidget(load_button)
        load_layout.addWidget(self.file_count_label)
        load_layout.addWidget(clear_button)
        
        load_group.setLayout(load_layout)
        tools_layout.addWidget(load_group)
        
        # 출력 옵션
        options_group = QGroupBox("출력 옵션")
        options_layout = QHBoxLayout()
        
        self.include_memory = QCheckBox("메모리 정보 포함")
        self.include_memory.setChecked(True)
        
        self.include_cpu = QCheckBox("CPU 정보 포함")
        self.include_cpu.setChecked(True)
        
        self.include_uptime = QCheckBox("가동시간 포함") 
        self.include_uptime.setChecked(True)
        
        options_layout.addWidget(self.include_memory)
        options_layout.addWidget(self.include_cpu)
        options_layout.addWidget(self.include_uptime)
        
        options_group.setLayout(options_layout)
        tools_layout.addWidget(options_group)
        
        # 결과 저장 버튼
        save_group = QGroupBox("보고서 저장")
        save_layout = QHBoxLayout()
        
        save_excel_button = QPushButton("엑셀로 저장")
        save_excel_button.setIcon(QIcon("icons/excel.png"))
        save_excel_button.clicked.connect(self.save_to_excel)
        
        save_csv_button = QPushButton("CSV로 저장")
        save_csv_button.setIcon(QIcon("icons/csv.png"))
        save_csv_button.clicked.connect(self.save_to_csv)
        
        save_layout.addWidget(save_excel_button)
        save_layout.addWidget(save_csv_button)
        
        save_group.setLayout(save_layout)
        tools_layout.addWidget(save_group)

        # 결과 테이블
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(10)
        self.result_table.setHorizontalHeaderLabels([
            "파일명", "호스트명", "모델", "IOS 버전", "SW 버전", 
            "CPU 사용률", "메모리 총량", "메모리 사용", "메모리 사용률", "가동시간"
        ])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSortingEnabled(True)
        

        self.result_table.setStyleSheet("""
            QTableWidget {
                background-color: #2A2A2A;
                gridline-color: #444444;
                color: #E0E0E0;
                border: 1px solid #555555;
            }
            QTableWidget::item {
                color: #E0E0E0;
                background-color: #2A2A2A;
            }
            QTableWidget::item:selected {
                background-color: #4B6EAF;
                color: white;
            }
            QTableWidget::item:alternate {
                background-color: #333333;  /* 짝수 행의 배경색 */
            }
            QHeaderView::section {
                background-color: #3D3D3D;
                color: white;
                font-weight: bold;
                border: 1px solid #555555;
                padding: 4px;
            }
        """)



        # 상세 정보 표시 영역
        details_group = QGroupBox("상세 정보")
        details_layout = QVBoxLayout()
        
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        self.details_text.setStyleSheet("font-family: Consolas, monospace;")
        
        details_layout.addWidget(self.details_text)
        details_group.setLayout(details_layout)
        
        # 분할 영역 생성
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.result_table)
        splitter.addWidget(details_group)
        splitter.setSizes([300, 200])  # 초기 크기 비율 설정
        
        # 시그널 연결
        self.result_table.itemClicked.connect(self.show_device_details)

        # 레이아웃에 위젯 추가
        layout.addLayout(tools_layout)
        layout.addWidget(splitter)
        
        self.setLayout(layout)

    def load_txt_files(self):
        """텍스트 파일을 불러와 파싱"""
        files, _ = QFileDialog.getOpenFileNames(self, "텍스트 파일 불러오기", "", "Text Files (*.txt);;All Files (*)")
        if not files:
            return
            
        self.details_text.append(f"[INFO] {len(files)}개 파일을 불러왔습니다:")
        for file in files:
            self.details_text.append(f" - {file}")
        
        self.parse_txt_files(files)
        self.file_count_label.setText(f"로드된 파일: {len(self.device_data)}개")
        logging.info(f"[INFO] 텍스트 파일 불러오기: {len(files)}개")

    def parse_txt_files(self, files):
        """텍스트 파일에서 장비 데이터 파싱"""
        for file in files:
            try:
                with open(file, "r", encoding="utf-8", errors="replace") as f:
                    content = f.read()

                # 파일명 추출
                filename = os.path.basename(file)
                filename_without_ext = os.path.splitext(filename)[0]  # 확장자 제거

                # 기존 파싱 함수 사용
                run_hostname = self.parse_run_hostname(content)
                ios_version, sw_version, model = self.parse_show_version(content)
                total_mem, used_mem, free_mem, memory_usage = self.parse_memory(content)
                cpu_usage = self.parse_cpu(content)
                uptime = self.parse_uptime(content)

                # 데이터 저장
                device_info = {
                    "filename": filename_without_ext,  # 파일명 저장
                    "hostname": run_hostname,
                    "cpu": cpu_usage,
                    "memory_total": total_mem,
                    "memory_used": used_mem,
                    "memory_free": free_mem,
                    "memory_usage": memory_usage,
                    "uptime": uptime,
                    "ios_version": ios_version,
                    "sw_version": sw_version,
                    "model": model,
                    "raw_content": content  # 원본 데이터 저장
                }
                self.device_data.append(device_info)
                
                # 테이블에 행 추가
                self.add_device_to_table(device_info)
                
            except Exception as e:
                self.details_text.append(f"[ERROR] {file}: 데이터 파싱 실패 - {e}")
                logging.error(f"[ERROR] {file}: Nexus 데이터 파싱 실패 - {e}")

    def add_device_to_table(self, device):
        """장치 정보를 테이블에 추가"""
        row_position = self.result_table.rowCount()
        self.result_table.insertRow(row_position)
        
        # 셀 데이터 설정
        self.result_table.setItem(row_position, 0, QTableWidgetItem(device["filename"]))
        self.result_table.setItem(row_position, 1, QTableWidgetItem(device["hostname"]))
        self.result_table.setItem(row_position, 2, QTableWidgetItem(device["model"]))
        self.result_table.setItem(row_position, 3, QTableWidgetItem(device["nxos_version"]))
        self.result_table.setItem(row_position, 4, QTableWidgetItem(device["cpu"]))
        self.result_table.setItem(row_position, 5, QTableWidgetItem(str(device["memory_total"])))
        self.result_table.setItem(row_position, 6, QTableWidgetItem(str(device["memory_used"])))
        self.result_table.setItem(row_position, 7, QTableWidgetItem(device["memory_usage"]))
        self.result_table.setItem(row_position, 8, QTableWidgetItem(device["uptime"]))
        self.result_table.setItem(row_position, 9, QTableWidgetItem(device["last_reboot"]))

    def show_device_details(self, item):
        """선택한 장치의 상세 정보 표시"""
        row = item.row()
        filename = self.result_table.item(row, 0).text()
        
        # 해당 파일명의 장치 찾기
        device = next((d for d in self.device_data if d["filename"] == filename), None)
        if not device:
            return
            
         # 상세 정보 표시
        self.details_text.clear()
        self.details_text.append(f"장치 상세정보: {filename} ({device['hostname']})")
        self.details_text.append(f"모델: {device['model']}")
        self.details_text.append(f"NX-OS 버전: {device['nxos_version']}")
        self.details_text.append(f"CPU 사용률: {device['cpu']}")
        self.details_text.append(f"메모리: 총 {device['memory_total']}, 사용 {device['memory_used']}, 여유 {device['memory_free']}, 사용률 {device['memory_usage']}")
        self.details_text.append(f"가동시간: {device['uptime']}")
        self.details_text.append(f"마지막 재부팅: {device['last_reboot']}")
        self.details_text.append("\n--- 원본 데이터 일부 ---")
        
        # 원본 데이터의 일부분만 표시 (너무 길 수 있으므로)
        content_preview = device.get("raw_content", "")
        if content_preview:
            # 앞부분 2000자 정도만 표시
            preview_length = min(2000, len(content_preview))
            self.details_text.append(content_preview[:preview_length])
            if preview_length < len(content_preview):
                self.details_text.append("... (더 많은 데이터 생략) ...")

    def clear_data(self):
        """데이터 초기화"""
        if not self.device_data:
            return
            
        reply = QMessageBox.question(
            self, "데이터 초기화", 
            "모든 로드된 데이터를 초기화하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.device_data.clear()
            self.result_table.setRowCount(0)
            self.details_text.clear()
            self.file_count_label.setText("로드된 파일: 0개")
            logging.info("[INFO] Nexus 데이터 초기화됨")

    def save_to_excel(self):
        """파싱된 데이터를 엑셀로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "엑셀 파일로 저장", "NX-OS_Report.xlsx", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return

            # Excel 워크북 생성
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Nexus Report"

            # 헤더 추가
            headers = ["파일명", "호스트명", "모델", "NX-OS 버전"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU 사용률")
            if self.include_memory.isChecked():
                headers.extend(["메모리 총량", "메모리 사용", "메모리 여유", "메모리 사용률"])
            if self.include_uptime.isChecked():
                headers.extend(["가동시간", "마지막 재부팅"])
                
            sheet.append(headers)

            # 데이터 추가
            for device in self.device_data:
                row_data = [
                    device["filename"],
                    device["hostname"],
                    device["model"],
                    device["nxos_version"]
                ]
                
                # 옵션에 따라 데이터 추가
                if self.include_cpu.isChecked():
                    row_data.append(device["cpu"])
                if self.include_memory.isChecked():
                    row_data.extend([
                        device["memory_total"],
                        device["memory_used"],
                        device["memory_free"],
                        device["memory_usage"]
                    ])
                if self.include_uptime.isChecked():
                    row_data.extend([
                        device["uptime"],
                        device["last_reboot"]
                    ])
                    
                sheet.append(row_data)

            # 파일 저장
            workbook.save(file_path)
            logging.info(f"[INFO] Nexus 엑셀 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"Nexus 보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"엑셀 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] Nexus 엑셀 저장 오류: {e}")

    def save_to_csv(self):
        """파싱된 데이터를 CSV로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "CSV 파일로 저장", "NX-OS_Report.csv", "CSV Files (*.csv)"
            )
            
            if not file_path:
                return

            # 헤더 준비
            headers = ["Filename", "Hostname", "Model", "NXOS_Version"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU_Usage")
            if self.include_memory.isChecked():
                headers.extend(["Memory_Total", "Memory_Used", "Memory_Free", "Memory_Usage"])
            if self.include_uptime.isChecked():
                headers.extend(["Uptime", "Last_Reboot"])

            # CSV 파일 작성
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                for device in self.device_data:
                    row_data = [
                        device["filename"],
                        device["hostname"],
                        device["model"],
                        device["nxos_version"]
                    ]
                    
                    # 옵션에 따라 데이터 추가
                    if self.include_cpu.isChecked():
                        row_data.append(device["cpu"])
                    if self.include_memory.isChecked():
                        row_data.extend([
                            device["memory_total"],
                            device["memory_used"],
                            device["memory_free"],
                            device["memory_usage"]
                        ])
                    if self.include_uptime.isChecked():
                        row_data.extend([
                            device["uptime"],
                            device["last_reboot"]
                        ])
                        
                    writer.writerow(row_data)

            logging.info(f"[INFO] Nexus CSV 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"Nexus CSV 보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"CSV 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] Nexus CSV 저장 오류: {e}")

    # Nexus 파싱 메서드들
    def parse_show_version(self, content):
        """NX-OS 호스트명, 버전 및 모델 파싱"""
        hostname_match = re.search(r"Device name:\s+(\S+)", content)
        hostname = hostname_match.group(1) if hostname_match else "N/A"
        
        nxos_version_match = re.search(r"NXOS: version\s+([\d\.\(\)A-Z]+)", content, re.IGNORECASE)
        if not nxos_version_match:
            nxos_version_match = re.search(r"System version:\s+([\d\.\(\)A-Z]+)", content, re.IGNORECASE)
        nxos_version = nxos_version_match.group(1) if nxos_version_match else "N/A"
        
        model_match = re.search(r"Hardware\s+:\s+cisco\s+Nexus\d+\s+(\S+)", content, re.IGNORECASE)
        if not model_match:
            model_match = re.search(r"\b(N\d[KX]?-\S+)\b", content)
        model = model_match.group(1) if model_match else "N/A"
        
        return hostname, nxos_version, model
        
    def parse_system_resources(self, content):
        """CPU, 메모리, 가동시간 파싱"""
        # CPU 사용률 파싱
        cpu_match = re.search(r"CPU states\s+:\s+([\d\.]+)% user,\s+([\d\.]+)% kernel,\s+([\d\.]+)% idle", content)
        if cpu_match:
            user_cpu = float(cpu_match.group(1))
            kernel_cpu = float(cpu_match.group(2))
            total_cpu = round(user_cpu + kernel_cpu, 2)
            cpu_usage = f"{total_cpu}%"
        else:
            cpu_usage = "N/A"
        
        # 메모리 사용률 파싱
        mem_match = re.search(r"Memory usage:\s+(\d+)K total,\s+(\d+)K used,\s+(\d+)K free", content)
        if mem_match:
            total_mem = int(mem_match.group(1))
            used_mem = int(mem_match.group(2))
            free_mem = int(mem_match.group(3))
            usage_percentage = round((used_mem / total_mem) * 100, 2)
            memory_usage = f"{usage_percentage}%"
        else:
            total_mem, used_mem, free_mem, memory_usage = "N/A", "N/A", "N/A", "N/A"
        
        # 가동시간 파싱
        uptime_match = re.search(r"Kernel uptime is (\d+) day\(s\), (\d+) hour\(s\), (\d+) minute\(s\), (\d+) second\(s\)", content)
        if uptime_match:
            days, hours, minutes, seconds = uptime_match.groups()
            uptime = f"{days}일 {hours}시간 {minutes}분 {seconds}초"
        else:
            # Alternative uptime format
            uptime_match = re.search(r"(\S+) uptime is (\d+ years, )?(\d+ weeks, )?(\d+ days, )?(\d+ hours, )?(\d+ minutes)", content)
            if uptime_match:
                uptime_parts = [part for part in uptime_match.groups()[1:] if part]
                uptime = ''.join(uptime_parts).strip()
            else:
                uptime = "N/A"
        
        return cpu_usage, total_mem, used_mem, free_mem, memory_usage, uptime
    
    def parse_host_from_filename(self, file_path):
        """파일명에서 호스트명 추출"""
        filename = os.path.basename(file_path)
        host = os.path.splitext(filename)[0]  # 확장자 제거
        # 간단한 호스트명 검증 (예: CNS22_3F_MDF_BB_2)
        if re.match(r'^[\w\-\.]+$', host):
            return host
        return "N/A"

    def add_device_to_table(self, device):
        """장치 정보를 테이블에 추가"""
        row_position = self.result_table.rowCount()
        self.result_table.insertRow(row_position)
        
        # 셀 데이터 설정
        self.result_table.setItem(row_position, 0, QTableWidgetItem(device["filename"]))
        self.result_table.setItem(row_position, 1, QTableWidgetItem(device["hostname"]))
        self.result_table.setItem(row_position, 2, QTableWidgetItem(device["model"]))
        self.result_table.setItem(row_position, 3, QTableWidgetItem(device["ios_version"]))
        self.result_table.setItem(row_position, 4, QTableWidgetItem(device["sw_version"]))
        self.result_table.setItem(row_position, 5, QTableWidgetItem(device["cpu"]))
        self.result_table.setItem(row_position, 6, QTableWidgetItem(str(device["memory_total"])))
        self.result_table.setItem(row_position, 7, QTableWidgetItem(str(device["memory_used"])))
        self.result_table.setItem(row_position, 8, QTableWidgetItem(device["memory_usage"]))
        self.result_table.setItem(row_position, 9, QTableWidgetItem(device["uptime"]))

    def show_device_details(self, item):
        """선택한 장치의 상세 정보 표시"""
        row = item.row()
        filename = self.result_table.item(row, 0).text()
        
        # 해당 파일명의 장치 찾기
        device = next((d for d in self.device_data if d["filename"] == filename), None)
        if not device:
            return
            
        # 상세 정보 표시
        self.details_text.clear()
        self.details_text.append(f"장치 상세정보: {filename} ({device['hostname']})")
        self.details_text.append(f"모델: {device['model']}")
        self.details_text.append(f"IOS 버전: {device['ios_version']}")
        self.details_text.append(f"SW 버전: {device['sw_version']}")
        self.details_text.append(f"CPU 사용률: {device['cpu']}")
        self.details_text.append(f"메모리: 총 {device['memory_total']}, 사용 {device['memory_used']}, 여유 {device['memory_free']}, 사용률 {device['memory_usage']}")
        self.details_text.append(f"가동시간: {device['uptime']}")
        self.details_text.append("\n--- 원본 데이터 일부 ---")
        
        # 원본 데이터의 일부분만 표시 (너무 길 수 있으므로)
        content_preview = device.get("raw_content", "")
        if content_preview:
            # 앞부분 2000자 정도만 표시
            preview_length = min(2000, len(content_preview))
            self.details_text.append(content_preview[:preview_length])
            if preview_length < len(content_preview):
                self.details_text.append("... (더 많은 데이터 생략) ...")

    def clear_data(self):
        """데이터 초기화"""
        if not self.device_data:
            return
            
        reply = QMessageBox.question(
            self, "데이터 초기화", 
            "모든 로드된 데이터를 초기화하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.device_data.clear()
            self.result_table.setRowCount(0)
            self.details_text.clear()
            self.file_count_label.setText("로드된 파일: 0개")
            logging.info("[INFO] 데이터 초기화됨")

    def save_to_excel(self):
        """파싱된 데이터를 엑셀로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "엑셀 파일로 저장", "IOS-XE_Report.xlsx", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return

            # Excel 워크북 생성
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Inspection Report"

            # 헤더 추가
            headers = ["파일명", "호스트명", "모델", "IOS 버전", "SW 버전"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU 사용률")
            if self.include_memory.isChecked():
                headers.extend(["메모리 총량", "메모리 사용", "메모리 여유", "메모리 사용률"])
            if self.include_uptime.isChecked():
                headers.append("가동시간")
                
            sheet.append(headers)

            # 데이터 추가
            for device in self.device_data:
                row_data = [
                    device["filename"],
                    device["hostname"],
                    device["model"],
                    device["ios_version"],
                    device["sw_version"]
                ]
                
                # 옵션에 따라 데이터 추가
                if self.include_cpu.isChecked():
                    row_data.append(device["cpu"])
                if self.include_memory.isChecked():
                    row_data.extend([
                        device["memory_total"],
                        device["memory_used"],
                        device["memory_free"],
                        device["memory_usage"]
                    ])
                if self.include_uptime.isChecked():
                    row_data.append(device["uptime"])
                    
                sheet.append(row_data)

            # 파일 저장
            workbook.save(file_path)
            logging.info(f"[INFO] 엑셀 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"엑셀 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] 엑셀 저장 오류: {e}")

    def save_to_csv(self):
        """파싱된 데이터를 CSV로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "CSV 파일로 저장", "IOS-XE_Report.csv", "CSV Files (*.csv)"
            )
            
            if not file_path:
                return

            # 헤더 준비
            headers = ["Filename", "Hostname", "Model", "IOS_Version", "SW_Version"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU_Usage")
            if self.include_memory.isChecked():
                headers.extend(["Memory_Total", "Memory_Used", "Memory_Free", "Memory_Usage"])
            if self.include_uptime.isChecked():
                headers.append("Uptime")

            # CSV 파일 작성
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                for device in self.device_data:
                    row_data = [
                        device["filename"],
                        device["hostname"],
                        device["model"],
                        device["ios_version"],
                        device["sw_version"]
                    ]
                    
                    # 옵션에 따라 데이터 추가
                    if self.include_cpu.isChecked():
                        row_data.append(device["cpu"])
                    if self.include_memory.isChecked():
                        row_data.extend([
                            device["memory_total"],
                            device["memory_used"],
                            device["memory_free"],
                            device["memory_usage"]
                        ])
                    if self.include_uptime.isChecked():
                        row_data.append(device["uptime"])
                        
                    writer.writerow(row_data)

            logging.info(f"[INFO] CSV 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"CSV 보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"CSV 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] CSV 저장 오류: {e}")

    # 기존 파싱 메서드들
    def parse_run_hostname(self, content):
        """설정에서 호스트명 파싱"""
        match = re.search(r"^hostname\s+(\S+)", content, re.MULTILINE)
        if match:
            return match.group(1)
        return "N/A"

    def parse_show_version(self, content):
        """버전 정보 및 모델 파싱"""
        ios_version_match = re.search(r"Version\s+([\d\.\(\)A-Z]+)", content)
        ios_version = ios_version_match.group(1) if ios_version_match else "N/A"
        
        sw_version = ios_version  # SW 버전은 IOS 버전과 동일하게 설정
        
        model_match = re.search(r"Model number\s+:\s+(\S+)", content, re.IGNORECASE)
        if model_match:
            model = model_match.group(1)
        else:
            specific_model_match = re.search(r"\bC\d{4,}\b", content)
            if specific_model_match:
                model = specific_model_match.group(0)
            else:
                model = "N/A"
                
        return ios_version, sw_version, model

    def parse_memory(self, content):
        """메모리 정보 파싱"""
        match = re.search(r"Processor Pool Total:\s+(\d+)\s+Used:\s+(\d+)\s+Free:\s+(\d+)", content, re.DOTALL)
        if match:
            total = int(match.group(1))
            used = int(match.group(2))
            free = int(match.group(3))
            usage_percentage = round((used / total) * 100, 2)
            return total, used, free, f"{usage_percentage}%"

        match = re.search(r"System memory\s+:\s+(\d+)K total,\s+(\d+)K used,\s+(\d+)K free", content)
        if match:
            total = int(match.group(1))
            used = int(match.group(2))
            free = int(match.group(3))
            usage_percentage = round((used / total) * 100, 2)
            return total, used, free, f"{usage_percentage}%"
            
        return "N/A", "N/A", "N/A", "N/A"

    def parse_cpu(self, content):
        """CPU 사용률 파싱"""
        match = re.search(r"CPU utilization for five seconds:.*?(\d+)%", content)
        if match:
            return match.group(1) + "%"
        return "N/A"

    def parse_uptime(self, content):
        """장비 가동시간 파싱"""
        match = re.search(r"uptime is (.+)", content)
        if match:
            return match.group(1).strip()
        return "N/A"


class EnhancedNexusReportGenerator(QWidget):
    def __init__(self):
        super().__init__()
        self.device_data = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 상단 도구 모음
        tools_layout = QHBoxLayout()
        
        # 파일 로드 섹션
        load_group = QGroupBox("데이터 소스")
        load_layout = QHBoxLayout()
        
        load_button = QPushButton("파일 불러오기")
        load_button.setIcon(QIcon("icons/open.png"))
        load_button.clicked.connect(self.load_txt_files)
        
        self.file_count_label = QLabel("로드된 파일: 0개")
        
        clear_button = QPushButton("초기화")
        clear_button.setIcon(QIcon("icons/clear.png"))
        clear_button.clicked.connect(self.clear_data)
        
        load_layout.addWidget(load_button)
        load_layout.addWidget(self.file_count_label)
        load_layout.addWidget(clear_button)
        
        load_group.setLayout(load_layout)
        tools_layout.addWidget(load_group)
        
        # 출력 옵션
        options_group = QGroupBox("출력 옵션")
        options_layout = QHBoxLayout()
        
        self.include_memory = QCheckBox("메모리 정보 포함")
        self.include_memory.setChecked(True)
        
        self.include_cpu = QCheckBox("CPU 정보 포함")
        self.include_cpu.setChecked(True)
        
        self.include_uptime = QCheckBox("가동시간 포함") 
        self.include_uptime.setChecked(True)
        
        options_layout.addWidget(self.include_memory)
        options_layout.addWidget(self.include_cpu)
        options_layout.addWidget(self.include_uptime)
        
        options_group.setLayout(options_layout)
        tools_layout.addWidget(options_group)
        
        # 결과 저장 버튼
        save_group = QGroupBox("보고서 저장")
        save_layout = QHBoxLayout()
        
        save_excel_button = QPushButton("엑셀로 저장")
        save_excel_button.setIcon(QIcon("icons/excel.png"))
        save_excel_button.clicked.connect(self.save_to_excel)
        
        save_csv_button = QPushButton("CSV로 저장")
        save_csv_button.setIcon(QIcon("icons/csv.png"))
        save_csv_button.clicked.connect(self.save_to_csv)
        
        save_layout.addWidget(save_excel_button)
        save_layout.addWidget(save_csv_button)
        
        save_group.setLayout(save_layout)
        tools_layout.addWidget(save_group)

        # 결과 테이블
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(10)
        self.result_table.setHorizontalHeaderLabels([
            "파일명", "호스트명", "모델", "NX-OS 버전", 
            "CPU 사용률", "메모리 총량", "메모리 사용", "메모리 사용률", 
            "가동시간", "마지막 재부팅"
        ])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSortingEnabled(True)


        self.result_table.setStyleSheet("""
            QTableWidget {
                background-color: #2A2A2A;
                gridline-color: #444444;
                color: #E0E0E0;
                border: 1px solid #555555;
            }
            QTableWidget::item {
                color: #E0E0E0;
                background-color: #2A2A2A;
            }
            QTableWidget::item:selected {
                background-color: #4B6EAF;
                color: white;
            }
            QTableWidget::item:alternate {
                background-color: #333333;  /* 짝수 행의 배경색 */
            }
            QHeaderView::section {
                background-color: #3D3D3D;
                color: white;
                font-weight: bold;
                border: 1px solid #555555;
                padding: 4px;
            }
        """)
        
        # 상세 정보 표시 영역
        details_group = QGroupBox("상세 정보")
        details_layout = QVBoxLayout()
        
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        self.details_text.setStyleSheet("font-family: Consolas, monospace;")
        
        details_layout.addWidget(self.details_text)
        details_group.setLayout(details_layout)
        
        # 분할 영역 생성
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.result_table)
        splitter.addWidget(details_group)
        splitter.setSizes([300, 200])  # 초기 크기 비율 설정
        
        # 시그널 연결
        self.result_table.itemClicked.connect(self.show_device_details)

        # 레이아웃에 위젯 추가
        layout.addLayout(tools_layout)
        layout.addWidget(splitter)
        
        self.setLayout(layout)

    def load_txt_files(self):
        """Nexus 텍스트 파일 불러오기"""
        files, _ = QFileDialog.getOpenFileNames(self, "Nexus 텍스트 파일 불러오기", "", "Text Files (*.txt);;All Files (*)")
        if not files:
            return
            
        self.details_text.append(f"[INFO] {len(files)}개 파일을 불러왔습니다:")
        for file in files:
            self.details_text.append(f" - {file}")
        
        self.parse_txt_files(files)
        self.file_count_label.setText(f"로드된 파일: {len(self.device_data)}개")
        logging.info(f"[INFO] Nexus 텍스트 파일 불러오기: {len(files)}개")

    def parse_txt_files(self, files):
        """Nexus 장비 데이터 파싱"""
        for file in files:
            try:
                with open(file, "r", encoding="utf-8", errors="replace") as f:
                    content = f.read()

                # 파일명 추출
                filename = os.path.basename(file)
                filename_without_ext = os.path.splitext(filename)[0]  # 확장자 제거

                # Nexus 데이터 파싱
                hostname, nxos_version, model = self.parse_show_version(content)
                cpu_usage, total_mem, used_mem, free_mem, memory_usage, uptime = self.parse_system_resources(content)
                host_from_filename = self.parse_host_from_filename(file)
                last_reboot = self.parse_last_reboot(content)

                # 데이터 저장
                device_info = {
                    "filename": filename_without_ext,  # 파일명 저장
                    "hostname": hostname if hostname != "N/A" else host_from_filename,
                    "cpu": cpu_usage,
                    "memory_total": total_mem,
                    "memory_used": used_mem,
                    "memory_free": free_mem,
                    "memory_usage": memory_usage,
                    "uptime": uptime,
                    "nxos_version": nxos_version,
                    "model": model,
                    "last_reboot": last_reboot,
                    "raw_content": content  # 원본 데이터 저장
                }
                self.device_data.append(device_info)
                
                # 테이블에 행 추가
                self.add_device_to_table(device_info)
                
            except Exception as e:
                self.details_text.append(f"[ERROR] {file}: 데이터 파싱 실패 - {e}")
                logging.error(f"[ERROR] {file}: Nexus 데이터 파싱 실패 - {e}")

    def parse_show_version(self, content):
        """NX-OS 호스트명, 버전 및 모델 파싱"""
        hostname_match = re.search(r"Device name:\s+(\S+)", content)
        hostname = hostname_match.group(1) if hostname_match else "N/A"
        
        nxos_version_match = re.search(r"NXOS: version\s+([\d\.\(\)A-Z]+)", content, re.IGNORECASE)
        if not nxos_version_match:
            nxos_version_match = re.search(r"System version:\s+([\d\.\(\)A-Z]+)", content, re.IGNORECASE)
        nxos_version = nxos_version_match.group(1) if nxos_version_match else "N/A"
        
        model_match = re.search(r"Hardware\s+:\s+cisco\s+Nexus\d+\s+(\S+)", content, re.IGNORECASE)
        if not model_match:
            model_match = re.search(r"\b(N\d[KX]?-\S+)\b", content)
        model = model_match.group(1) if model_match else "N/A"
        
        return hostname, nxos_version, model
    
    def parse_system_resources(self, content):
        """CPU, 메모리, 가동시간 파싱"""
        # CPU 사용률 파싱
        cpu_match = re.search(r"CPU states\s+:\s+([\d\.]+)% user,\s+([\d\.]+)% kernel,\s+([\d\.]+)% idle", content)
        if cpu_match:
            user_cpu = float(cpu_match.group(1))
            kernel_cpu = float(cpu_match.group(2))
            total_cpu = round(user_cpu + kernel_cpu, 2)
            cpu_usage = f"{total_cpu}%"
        else:
            cpu_usage = "N/A"
        
        # 메모리 사용률 파싱
        mem_match = re.search(r"Memory usage:\s+(\d+)K total,\s+(\d+)K used,\s+(\d+)K free", content)
        if mem_match:
            total_mem = int(mem_match.group(1))
            used_mem = int(mem_match.group(2))
            free_mem = int(mem_match.group(3))
            usage_percentage = round((used_mem / total_mem) * 100, 2)
            memory_usage = f"{usage_percentage}%"
        else:
            total_mem, used_mem, free_mem, memory_usage = "N/A", "N/A", "N/A", "N/A"
        
        # 가동시간 파싱
        uptime_match = re.search(r"Kernel uptime is (\d+) day\(s\), (\d+) hour\(s\), (\d+) minute\(s\), (\d+) second\(s\)", content)
        if uptime_match:
            days, hours, minutes, seconds = uptime_match.groups()
            uptime = f"{days}일 {hours}시간 {minutes}분 {seconds}초"
        else:
            # Alternative uptime format
            uptime_match = re.search(r"(\S+) uptime is (\d+ years, )?(\d+ weeks, )?(\d+ days, )?(\d+ hours, )?(\d+ minutes)", content)
            if uptime_match:
                uptime_parts = [part for part in uptime_match.groups()[1:] if part]
                uptime = ''.join(uptime_parts).strip()
            else:
                uptime = "N/A"
        
        return cpu_usage, total_mem, used_mem, free_mem, memory_usage, uptime
    
    def parse_host_from_filename(self, file_path):
        """파일명에서 호스트명 추출"""
        filename = os.path.basename(file_path)
        host = os.path.splitext(filename)[0]  # 확장자 제거
        # 간단한 호스트명 검증 (예: CNS22_3F_MDF_BB_2)
        if re.match(r'^[\w\-\.]+$', host):
            return host
        return "N/A"
        
    def parse_last_reboot(self, content):
        """마지막 재부팅 시간 추출"""
        reboot_match = re.search(r"Last reset at\s+(.+?)(\s+Reason:|\n)", content)
        if reboot_match:
            return reboot_match.group(1).strip()
        return "N/A"

    def add_device_to_table(self, device):
        """장치 정보를 테이블에 추가"""
        row_position = self.result_table.rowCount()
        self.result_table.insertRow(row_position)
        
        # 셀 데이터 설정
        self.result_table.setItem(row_position, 0, QTableWidgetItem(device["filename"]))
        self.result_table.setItem(row_position, 1, QTableWidgetItem(device["hostname"]))
        self.result_table.setItem(row_position, 2, QTableWidgetItem(device["model"]))
        self.result_table.setItem(row_position, 3, QTableWidgetItem(device["nxos_version"]))
        self.result_table.setItem(row_position, 4, QTableWidgetItem(device["cpu"]))
        self.result_table.setItem(row_position, 5, QTableWidgetItem(str(device["memory_total"])))
        self.result_table.setItem(row_position, 6, QTableWidgetItem(str(device["memory_used"])))
        self.result_table.setItem(row_position, 7, QTableWidgetItem(device["memory_usage"]))
        self.result_table.setItem(row_position, 8, QTableWidgetItem(device["uptime"]))
        self.result_table.setItem(row_position, 9, QTableWidgetItem(device["last_reboot"]))

    def show_device_details(self, item):
        """선택한 장치의 상세 정보 표시"""
        row = item.row()
        filename = self.result_table.item(row, 0).text()
        
        # 해당 파일명의 장치 찾기
        device = next((d for d in self.device_data if d["filename"] == filename), None)
        if not device:
            return
            
        # 상세 정보 표시
        self.details_text.clear()
        self.details_text.append(f"장치 상세정보: {filename} ({device['hostname']})")
        self.details_text.append(f"모델: {device['model']}")
        self.details_text.append(f"NX-OS 버전: {device['nxos_version']}")
        self.details_text.append(f"CPU 사용률: {device['cpu']}")
        self.details_text.append(f"메모리: 총 {device['memory_total']}, 사용 {device['memory_used']}, 여유 {device['memory_free']}, 사용률 {device['memory_usage']}")
        self.details_text.append(f"가동시간: {device['uptime']}")
        self.details_text.append(f"마지막 재부팅: {device['last_reboot']}")
        self.details_text.append("\n--- 원본 데이터 일부 ---")
        
        # 원본 데이터의 일부분만 표시 (너무 길 수 있으므로)
        content_preview = device.get("raw_content", "")
        if content_preview:
            # 앞부분 2000자 정도만 표시
            preview_length = min(2000, len(content_preview))
            self.details_text.append(content_preview[:preview_length])
            if preview_length < len(content_preview):
                self.details_text.append("... (더 많은 데이터 생략) ...")

    def clear_data(self):
        """데이터 초기화"""
        if not self.device_data:
            return
            
        reply = QMessageBox.question(
            self, "데이터 초기화", 
            "모든 로드된 데이터를 초기화하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.device_data.clear()
            self.result_table.setRowCount(0)
            self.details_text.clear()
            self.file_count_label.setText("로드된 파일: 0개")
            logging.info("[INFO] Nexus 데이터 초기화됨")

    def save_to_excel(self):
        """파싱된 데이터를 엑셀로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "엑셀 파일로 저장", "NX-OS_Report.xlsx", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return

            # Excel 워크북 생성
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Nexus Report"

            # 헤더 추가
            headers = ["파일명", "호스트명", "모델", "NX-OS 버전"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU 사용률")
            if self.include_memory.isChecked():
                headers.extend(["메모리 총량", "메모리 사용", "메모리 여유", "메모리 사용률"])
            if self.include_uptime.isChecked():
                headers.extend(["가동시간", "마지막 재부팅"])
                
            sheet.append(headers)

            # 데이터 추가
            for device in self.device_data:
                row_data = [
                    device["filename"],
                    device["hostname"],
                    device["model"],
                    device["nxos_version"]
                ]
                
                # 옵션에 따라 데이터 추가
                if self.include_cpu.isChecked():
                    row_data.append(device["cpu"])
                if self.include_memory.isChecked():
                    row_data.extend([
                        device["memory_total"],
                        device["memory_used"],
                        device["memory_free"],
                        device["memory_usage"]
                    ])
                if self.include_uptime.isChecked():
                    row_data.extend([
                        device["uptime"],
                        device["last_reboot"]
                    ])
                    
                sheet.append(row_data)

            # 파일 저장
            workbook.save(file_path)
            logging.info(f"[INFO] Nexus 엑셀 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"Nexus 보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"엑셀 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] Nexus 엑셀 저장 오류: {e}")

    def save_to_csv(self):
        """파싱된 데이터를 CSV로 저장"""
        if not self.device_data:
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다. 먼저 데이터를 수집하세요.")
            return

        try:
            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "CSV 파일로 저장", "NX-OS_Report.csv", "CSV Files (*.csv)"
            )
            
            if not file_path:
                return

            # 헤더 준비
            headers = ["Filename", "Hostname", "Model", "NXOS_Version"]
            
            # 옵션에 따라 헤더 추가
            if self.include_cpu.isChecked():
                headers.append("CPU_Usage")
            if self.include_memory.isChecked():
                headers.extend(["Memory_Total", "Memory_Used", "Memory_Free", "Memory_Usage"])
            if self.include_uptime.isChecked():
                headers.extend(["Uptime", "Last_Reboot"])

            # CSV 파일 작성
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                for device in self.device_data:
                    row_data = [
                        device["filename"],
                        device["hostname"],
                        device["model"],
                        device["nxos_version"]
                    ]
                    
                    # 옵션에 따라 데이터 추가
                    if self.include_cpu.isChecked():
                        row_data.append(device["cpu"])
                    if self.include_memory.isChecked():
                        row_data.extend([
                            device["memory_total"],
                            device["memory_used"],
                            device["memory_free"],
                            device["memory_usage"]
                        ])
                    if self.include_uptime.isChecked():
                        row_data.extend([
                            device["uptime"],
                            device["last_reboot"]
                        ])
                        
                    writer.writerow(row_data)

            logging.info(f"[INFO] Nexus CSV 파일 저장됨: {file_path}")
            QMessageBox.information(self, "저장 완료", f"Nexus CSV 보고서가 저장되었습니다:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"CSV 파일 저장 중 오류 발생: {e}")
            logging.error(f"[ERROR] Nexus CSV 저장 오류: {e}")